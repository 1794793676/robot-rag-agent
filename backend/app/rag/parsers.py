"""Text extraction for txt, docx, xls/xlsx, and text-layer PDFs."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re

from docx import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd


class DocumentParseError(ValueError):
    """Raised when a supported file cannot provide usable text."""


@dataclass(slots=True)
class ParsedSection:
    text: str
    page: int | None = None
    heading: str | None = None


def _parse_txt(path: Path) -> list[ParsedSection]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise DocumentParseError("无法识别 TXT 文件编码，请使用 UTF-8、GBK 或 GB18030")
    if not text.strip():
        raise DocumentParseError("文档没有可用文本")
    return [ParsedSection(text=text.strip())]


def _table_to_markdown(table) -> str:
    rows = [[cell.text.strip().replace("\n", " ") for cell in row.cells] for row in table.rows]
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = "| " + " | ".join(normalized[0]) + " |"
    separator = "| " + " | ".join(["---"] * width) + " |"
    body = ["| " + " | ".join(row) + " |" for row in normalized[1:]]
    return "\n".join([header, separator, *body])


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip().replace("\n", " ")


def _trim_row(row: list[str]) -> list[str]:
    while row and not row[-1]:
        row.pop()
    return row


def _rows_to_markdown(rows: list[list[str]]) -> str:
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = "| " + " | ".join(normalized[0]) + " |"
    separator = "| " + " | ".join(["---"] * width) + " |"
    body = ["| " + " | ".join(row) + " |" for row in normalized[1:]]
    return "\n".join([header, separator, *body])


_CHINESE_NUMBERED_HEADING = re.compile(
    r"^(?:[一二三四五六七八九十百]+[、．.]|（[一二三四五六七八九十百]+）)\s*\S+$"
)
_WORD_HEADING_STYLE = re.compile(r"^(?:heading|标题)\s*(\d+)$")


def _heading_level(paragraph: Paragraph) -> int | None:
    style_name = (paragraph.style.name if paragraph.style else "").strip().lower()
    match = _WORD_HEADING_STYLE.match(style_name)
    if not match:
        return None
    return int(match.group(1))


def _is_heading(paragraph: Paragraph, text: str) -> bool:
    if _heading_level(paragraph) is not None:
        return True
    return (
        len(text) <= 20
        and text[-1] not in "。！？!?；;"
        and bool(_CHINESE_NUMBERED_HEADING.fullmatch(text))
    )


def _parse_docx(path: Path) -> list[ParsedSection]:
    try:
        document = DocxDocument(path)
    except Exception as exc:
        raise DocumentParseError(f"DOCX 解析失败：{exc}") from exc

    sections: list[ParsedSection] = []
    heading_stack: list[str] = []
    current_heading: str | None = None
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            paragraph = Paragraph(child, document)
            text = paragraph.text.strip()
            if not text:
                continue
            if _is_heading(paragraph, text):
                level = _heading_level(paragraph)
                if level is None:
                    heading_stack = [text]
                    current_heading = text
                else:
                    heading_stack = heading_stack[: max(0, level - 1)]
                    heading_stack.append(text)
                    current_heading = " / ".join(heading_stack)
                continue
        elif isinstance(child, CT_Tbl):
            text = _table_to_markdown(Table(child, document))
            if not text:
                continue
        else:
            continue
        sections.append(ParsedSection(text=text, heading=current_heading))

    if not sections:
        raise DocumentParseError("DOCX 文档没有可用文本")
    return sections


def _parse_xlsx(path: Path) -> list[ParsedSection]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentParseError(f"XLSX 解析失败：{exc}") from exc

    sections: list[ParsedSection] = []
    try:
        for sheet in workbook.worksheets:
            current_rows: list[tuple[int, list[str]]] = []

            def flush_rows() -> None:
                if not current_rows:
                    return
                start_row = current_rows[0][0]
                end_row = current_rows[-1][0]
                rows = [row for _, row in current_rows]
                text = _rows_to_markdown(rows)
                sections.append(
                    ParsedSection(
                        text=text,
                        heading=f"工作表 {sheet.title} / 行 {start_row}-{end_row}",
                    )
                )
                current_rows.clear()

            for row_index, raw_row in enumerate(
                sheet.iter_rows(values_only=True), start=1
            ):
                row = _trim_row([_cell_to_text(value) for value in raw_row])
                if not row or not any(row):
                    flush_rows()
                    continue
                current_rows.append((row_index, row))
            flush_rows()
    finally:
        workbook.close()

    if not sections:
        raise DocumentParseError("XLSX 文档没有可用文本")
    return sections


def _xls_value_to_text(value) -> str:
    if value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _cell_to_text(value)


def _parse_xls(path: Path) -> list[ParsedSection]:
    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise DocumentParseError(f"XLS 解析失败：{exc}") from exc

    sections: list[ParsedSection] = []
    try:
        for sheet in workbook.sheets():
            current_rows: list[tuple[int, list[str]]] = []

            def flush_rows() -> None:
                if not current_rows:
                    return
                start_row = current_rows[0][0]
                end_row = current_rows[-1][0]
                rows = [row for _, row in current_rows]
                sections.append(
                    ParsedSection(
                        text=_rows_to_markdown(rows),
                        heading=f"工作表 {sheet.name} / 行 {start_row}-{end_row}",
                    )
                )
                current_rows.clear()

            for zero_based_index in range(sheet.nrows):
                row_index = zero_based_index + 1
                row = _trim_row(
                    [_xls_value_to_text(value) for value in sheet.row_values(zero_based_index)]
                )
                if not row or not any(row):
                    flush_rows()
                    continue
                current_rows.append((row_index, row))
            flush_rows()
    finally:
        workbook.release_resources()

    if not sections:
        raise DocumentParseError("XLS 文档没有可用文本")
    return sections


def _parse_pdf(path: Path) -> list[ParsedSection]:
    try:
        reader = PdfReader(str(path))
        sections = [
            ParsedSection(text=text.strip(), page=index)
            for index, page in enumerate(reader.pages, start=1)
            if (text := (page.extract_text() or "")).strip()
        ]
    except Exception as exc:
        raise DocumentParseError(f"PDF 解析失败：{exc}") from exc
    total_chars = sum(len(section.text) for section in sections)
    if total_chars < 20:
        raise DocumentParseError(
            "当前版本不支持 OCR，请上传可复制文本的 PDF、docx 或 txt 文件。"
        )
    return sections


def parse_document(path: Path, file_type: str) -> list[ParsedSection]:
    """Parse a document into page-aware text sections."""

    parsers = {
        "txt": _parse_txt,
        "docx": _parse_docx,
        "xls": _parse_xls,
        "xlsx": _parse_xlsx,
        "pdf": _parse_pdf,
    }
    try:
        parser = parsers[file_type]
    except KeyError as exc:
        raise DocumentParseError("仅支持 txt、docx、xls、xlsx、pdf 文件") from exc
    return parser(path)
